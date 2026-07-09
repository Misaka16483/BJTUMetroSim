from __future__ import annotations

import base64
import json
import re
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.models import ValidationIssue, ValidationReport


JsonDict = dict[str, Any]
SENTINEL = 65535


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        if stripped == str(SENTINEL):
            return None
        return stripped
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if value == SENTINEL:
        return None
    return value


def to_int(value: Any) -> int | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        if value.lower().startswith("0x"):
            return int(value, 16)
        return int(float(value))
    return int(value)


def to_float(value: Any) -> float | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, str):
        if value.lower().startswith("0x"):
            return float(int(value, 16))
        return float(value)
    return float(value)


def cm_to_m(value: Any) -> float | None:
    numeric = to_float(value)
    return None if numeric is None else round(numeric / 100.0, 6)


def cmps_to_mps(value: Any) -> float | None:
    numeric = to_float(value)
    return None if numeric is None else round(numeric / 100.0, 6)


def parse_k_mileage(value: Any) -> float | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.fullmatch(r"[Kk]\s*(\d+)\s*\+\s*(\d+(?:\.\d+)?)", str(value))
    if not match:
        return None
    return int(match.group(1)) * 1000.0 + float(match.group(2))


class LineDataImporter:
    def import_file(self, source: str | Path) -> JsonDict:
        source_path = Path(source).resolve()
        workbook_path: Path | None = None
        temp_dir: tempfile.TemporaryDirectory[str] | None = None
        try:
            if source_path.suffix.lower() == ".xls":
                temp_dir = tempfile.TemporaryDirectory()
                workbook_path = Path(temp_dir.name) / "line_data.xlsx"
                self._convert_xls_to_xlsx(source_path, workbook_path)
            else:
                workbook_path = source_path
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
            try:
                tables = {name: self._read_table(workbook, name) for name in workbook.sheetnames}
            finally:
                workbook.close()
        finally:
            if temp_dir is not None:
                temp_dir.cleanup()

        line_map = self._build_line_map(source_path, tables)
        report = validate_line_map(line_map)
        line_map["validation"] = report.to_dict()
        return line_map

    def write_cache(self, line_map: JsonDict, cache_dir: str | Path) -> dict[str, Path]:
        cache_path = Path(cache_dir)
        cache_path.mkdir(parents=True, exist_ok=True)
        line_map_path = cache_path / "line_map.json"
        report_path = cache_path / "import_report.json"
        with line_map_path.open("w", encoding="utf-8") as handle:
            json.dump(line_map, handle, ensure_ascii=False, indent=2)
        with report_path.open("w", encoding="utf-8") as handle:
            json.dump(line_map.get("validation", {}), handle, ensure_ascii=False, indent=2)
        return {"line_map": line_map_path, "report": report_path}

    @staticmethod
    def _convert_xls_to_xlsx(source: Path, target: Path) -> None:
        source_b64 = base64.b64encode(str(source).encode("utf-8")).decode("ascii")
        target_b64 = base64.b64encode(str(target).encode("utf-8")).decode("ascii")
        script = f"""
$ErrorActionPreference = 'Stop'
$src = [System.Text.Encoding]::UTF8.GetString([Convert]::FromBase64String('{source_b64}'))
$dst = [System.Text.Encoding]::UTF8.GetString([Convert]::FromBase64String('{target_b64}'))
$tmp = [System.IO.Path]::ChangeExtension([System.IO.Path]::GetTempFileName(), '.xls')
$inputStream = [System.IO.File]::Open($src, [System.IO.FileMode]::Open, [System.IO.FileAccess]::Read, [System.IO.FileShare]::ReadWrite)
$outputStream = [System.IO.File]::Create($tmp)
try {{
  $inputStream.CopyTo($outputStream)
}} finally {{
  $outputStream.Close()
  $inputStream.Close()
}}
if (Test-Path $dst) {{ Remove-Item -LiteralPath $dst -Force }}
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$workbook = $null
try {{
  $workbook = $excel.Workbooks.Open($tmp, 0, $true)
  $workbook.SaveAs($dst, 51)
}} finally {{
  if ($workbook -ne $null) {{
    $workbook.Close($false)
    [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
  }}
  $excel.Quit()
  [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
  Remove-Item -LiteralPath $tmp -Force -ErrorAction SilentlyContinue
}}
"""
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout).strip()
            raise RuntimeError(f"Excel .xls conversion failed: {detail}")

    @staticmethod
    def _read_table(workbook: Any, sheet_name: str) -> list[JsonDict]:
        sheet = workbook[sheet_name]
        row_count = to_int(sheet.cell(2, 3).value) or 0
        col_count = to_int(sheet.cell(2, 4).value) or sheet.max_column
        headers = [
            str(sheet.cell(4, column).value or f"col_{column}").strip()
            for column in range(1, col_count + 1)
        ]
        rows: list[JsonDict] = []
        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=5,
                max_row=4 + row_count,
                min_col=1,
                max_col=col_count,
                values_only=True,
            ),
            start=5,
        ):
            normalized = [normalize_cell(value) for value in values]
            if all(value is None for value in normalized):
                continue
            raw = {
                headers[index]: normalized[index]
                for index in range(min(len(headers), len(normalized)))
            }
            raw["_rowNumber"] = row_number
            raw["_values"] = normalized
            rows.append(raw)
        return rows

    def _build_line_map(self, source: Path, tables: dict[str, list[JsonDict]]) -> JsonDict:
        return {
            "schemaVersion": "phase0.line-map.v1",
            "source": str(source),
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "counts": {name: len(rows) for name, rows in tables.items()},
            "points": [self._point(row) for row in tables.get("点表", [])],
            "segments": [self._segment(row) for row in tables.get("Seg表", [])],
            "switches": [self._switch(row) for row in tables.get("道岔表", [])],
            "signals": [self._signal(row) for row in tables.get("信号机表", [])],
            "stations": [self._station(row) for row in tables.get("车站表", [])],
            "platforms": [self._platform(row) for row in tables.get("站台表", [])],
            "balises": [self._balise(row) for row in tables.get("应答器表", [])],
            "gradients": [self._gradient(row) for row in tables.get("坡度表", [])],
            "speedRestrictions": [
                self._speed_restriction(row) for row in tables.get("静态限速表", [])
            ],
            "axleSections": [self._axle_section(row) for row in tables.get("计轴区段表", [])],
            "logicalSections": [
                self._logical_section(row) for row in tables.get("逻辑区段表", [])
            ],
            "protectionSections": [
                self._id_list_section(row, "axleSectionIds", 1, 2, 4)
                for row in tables.get("保护区段表", [])
            ],
            "pointApproachSections": [
                self._id_list_section(row, "axleSectionIds", 1, 2, 10)
                for row in tables.get("点式接近区段表", [])
            ],
            "cbtcApproachSections": [
                self._id_list_section(row, "logicalSectionIds", 1, 2, 10)
                for row in tables.get("CBTC接近区段表", [])
            ],
            "pointTriggerSections": [
                self._id_list_section(row, "axleSectionIds", 1, 2, 16)
                for row in tables.get("点式触发区段表", [])
            ],
            "cbtcTriggerSections": [
                self._id_list_section(row, "logicalSectionIds", 1, 2, 16)
                for row in tables.get("CBTC触发区段表", [])
            ],
            "routes": [self._route(row) for row in tables.get("进路表", [])],
        }

    @staticmethod
    def _values(row: JsonDict) -> list[Any]:
        return row["_values"]

    def _with_raw(self, row: JsonDict, data: JsonDict) -> JsonDict:
        data["raw"] = {key: value for key, value in row.items() if not key.startswith("_")}
        return data

    def _point(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "mileageLabel": values[2],
                "mileageM": parse_k_mileage(values[2]),
            },
        )

    def _segment(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "lengthCm": to_int(values[1]),
                "lengthM": cm_to_m(values[1]),
                "startEndpointType": to_int(values[2]),
                "startEndpointId": to_int(values[3]),
                "endEndpointType": to_int(values[4]),
                "endEndpointId": to_int(values[5]),
                "startForwardSegId": to_int(values[6]),
                "startDivergingSegId": to_int(values[7]),
                "endForwardSegId": to_int(values[8]),
                "endDivergingSegId": to_int(values[9]),
                "zcAreaId": to_int(values[10]),
                "atsAreaId": to_int(values[11]),
                "ciAreaId": to_int(values[12]),
                "hasSpeedInfo": to_int(values[18]),
                "hasGradientInfo": to_int(values[19]),
                "hasTunnelInfo": to_int(values[20]),
                "interoperabilityId": to_int(values[22]),
                "trackSectionProperty": values[23],
            },
        )

    def _switch(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "type": to_int(values[3]),
                "normalSegId": to_int(values[4]),
                "reverseSegId": to_int(values[5]),
                "frogSegId": to_int(values[6]),
            },
        )

    def _signal(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "type": to_int(values[2]),
                "attribute": values[3],
                "segmentId": to_int(values[4]),
                "offsetCm": to_int(values[5]),
                "offsetM": cm_to_m(values[5]),
                "direction": values[6],
                "aspectInfo": values[7],
                "interoperabilityId": to_int(values[8]),
            },
        )

    def _station(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(row, {"id": to_int(values[0]), "name": values[1]})

    def _platform(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "mileageLabel": values[1],
                "mileageM": parse_k_mileage(values[1]),
                "segmentId": to_int(values[2]),
                "direction": values[3],
                "triggerAxleSectionIds": self._take_ids(values, 4, 5, 6),
                "clearPassengerFlag": values[11],
                "interoperabilityId": to_int(values[12]),
                "offsetM": 0.0,
            },
        )

    def _balise(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "baliseId": values[1] if len(values) > 1 else None,
                "name": values[2] if len(values) > 2 else None,
                "segmentId": to_int(values[3]) if len(values) > 3 else None,
                "offsetCm": to_int(values[4]) if len(values) > 4 else None,
                "offsetM": cm_to_m(values[4]) if len(values) > 4 else None,
                "interoperabilityId": to_int(values[5]) if len(values) > 5 else None,
            },
        )

    def _gradient(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "startSegmentId": to_int(values[1]),
                "startOffsetCm": to_int(values[2]),
                "startOffsetM": cm_to_m(values[2]),
                "endSegmentId": to_int(values[3]),
                "endOffsetCm": to_int(values[4]),
                "endOffsetM": cm_to_m(values[4]),
                "startSwitchId": to_int(values[5]),
                "endSwitchId": to_int(values[8]),
                "slopePermille": to_int(values[11]),
                "direction": values[12],
                "verticalCurveRadius": to_int(values[13]),
            },
        )

    def _speed_restriction(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "segmentId": to_int(values[1]),
                "startOffsetCm": to_int(values[2]),
                "startOffsetM": cm_to_m(values[2]),
                "endOffsetCm": to_int(values[3]),
                "endOffsetM": cm_to_m(values[3]),
                "switchId": to_int(values[4]),
                "speedLimitCmps": to_int(values[5]),
                "speedLimitMps": cmps_to_mps(values[5]),
            },
        )

    def _axle_section(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "segmentIds": self._take_ids(values, 2, 3, 5),
            },
        )

    def _logical_section(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "startSegmentId": to_int(values[2]),
                "startOffsetCm": to_int(values[3]),
                "startOffsetM": cm_to_m(values[3]),
                "endSegmentId": to_int(values[4]),
                "endOffsetCm": to_int(values[5]),
                "endOffsetM": cm_to_m(values[5]),
            },
        )

    def _id_list_section(
        self,
        row: JsonDict,
        field_name: str,
        count_index: int,
        first_index: int,
        max_items: int,
    ) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                field_name: self._take_ids(values, count_index, first_index, max_items),
            },
        )

    def _route(self, row: JsonDict) -> JsonDict:
        values = self._values(row)
        return self._with_raw(
            row,
            {
                "id": to_int(values[0]),
                "name": values[1],
                "type": values[2],
                "startSignalId": to_int(values[3]),
                "endSignalId": to_int(values[4]),
                "axleSectionIds": self._take_ids(values, 5, 6, 20),
                "protectionSectionIds": self._take_ids(values, 26, 27, 5),
                "pointApproachSectionIds": self._take_ids(values, 32, 33, 5),
                "cbtcApproachSectionIds": self._take_ids(values, 38, 39, 5),
                "pointTriggerSectionIds": self._take_ids(values, 44, 45, 5),
                "cbtcTriggerSectionIds": self._take_ids(values, 50, 51, 5),
                "ciAreaId": to_int(values[56]),
            },
        )

    @staticmethod
    def _take_ids(values: list[Any], count_index: int, first_index: int, max_items: int) -> list[int]:
        count = to_int(values[count_index]) or 0
        selected = values[first_index : first_index + min(count, max_items)]
        ids = [to_int(value) for value in selected]
        return [item for item in ids if item is not None]


def validate_line_map(line_map: JsonDict) -> ValidationReport:
    counts = line_map.get("counts", {})
    report = ValidationReport(ok=True, counts=counts)

    def add_error(code: str, message: str, entity: str, entity_id: Any) -> None:
        report.errors.append(ValidationIssue("error", code, message, entity, entity_id))

    def ids(name: str) -> set[int]:
        return {
            int(item["id"])
            for item in line_map.get(name, [])
            if item.get("id") is not None
        }

    point_ids = ids("points")
    segment_ids = ids("segments")
    switch_ids = ids("switches")
    signal_ids = ids("signals")
    axle_ids = ids("axleSections")
    logical_ids = ids("logicalSections")
    protection_ids = ids("protectionSections")
    point_approach_ids = ids("pointApproachSections")
    cbtc_approach_ids = ids("cbtcApproachSections")
    point_trigger_ids = ids("pointTriggerSections")
    cbtc_trigger_ids = ids("cbtcTriggerSections")
    endpoint_ids = point_ids | switch_ids

    def check_ref(value: Any, target_ids: set[int], code: str, entity: str, entity_id: Any) -> None:
        if value is None:
            return
        if int(value) not in target_ids:
            add_error(code, f"missing reference id={value}", entity, entity_id)

    for segment in line_map.get("segments", []):
        entity_id = segment.get("id")
        if point_ids:
            check_ref(segment.get("startEndpointId"), endpoint_ids, "SEG_ENDPOINT", "segment", entity_id)
            check_ref(segment.get("endEndpointId"), endpoint_ids, "SEG_ENDPOINT", "segment", entity_id)
        for key in (
            "startForwardSegId",
            "startDivergingSegId",
            "endForwardSegId",
            "endDivergingSegId",
        ):
            check_ref(segment.get(key), segment_ids, "SEG_ADJACENT", "segment", entity_id)

    for signal in line_map.get("signals", []):
        check_ref(signal.get("segmentId"), segment_ids, "SIGNAL_SEGMENT", "signal", signal.get("id"))

    for platform in line_map.get("platforms", []):
        check_ref(platform.get("segmentId"), segment_ids, "PLATFORM_SEGMENT", "platform", platform.get("id"))
        for axle_id in platform.get("triggerAxleSectionIds", []):
            check_ref(axle_id, axle_ids, "PLATFORM_AXLE", "platform", platform.get("id"))

    for balise in line_map.get("balises", []):
        check_ref(balise.get("segmentId"), segment_ids, "BALISE_SEGMENT", "balise", balise.get("id"))

    for speed in line_map.get("speedRestrictions", []):
        check_ref(speed.get("segmentId"), segment_ids, "SPEED_SEGMENT", "speedRestriction", speed.get("id"))
        check_ref(speed.get("switchId"), switch_ids, "SPEED_SWITCH", "speedRestriction", speed.get("id"))

    for gradient in line_map.get("gradients", []):
        check_ref(gradient.get("startSegmentId"), segment_ids, "GRADIENT_SEGMENT", "gradient", gradient.get("id"))
        check_ref(gradient.get("endSegmentId"), segment_ids, "GRADIENT_SEGMENT", "gradient", gradient.get("id"))
        check_ref(gradient.get("startSwitchId"), switch_ids, "GRADIENT_SWITCH", "gradient", gradient.get("id"))
        check_ref(gradient.get("endSwitchId"), switch_ids, "GRADIENT_SWITCH", "gradient", gradient.get("id"))

    for axle in line_map.get("axleSections", []):
        for segment_id in axle.get("segmentIds", []):
            check_ref(segment_id, segment_ids, "AXLE_SEGMENT", "axleSection", axle.get("id"))

    for logical in line_map.get("logicalSections", []):
        check_ref(logical.get("startSegmentId"), segment_ids, "LOGICAL_SEGMENT", "logicalSection", logical.get("id"))
        check_ref(logical.get("endSegmentId"), segment_ids, "LOGICAL_SEGMENT", "logicalSection", logical.get("id"))

    for section in line_map.get("protectionSections", []):
        for axle_id in section.get("axleSectionIds", []):
            check_ref(axle_id, axle_ids, "PROTECTION_AXLE", "protectionSection", section.get("id"))
    for section in line_map.get("pointApproachSections", []):
        for axle_id in section.get("axleSectionIds", []):
            check_ref(axle_id, axle_ids, "POINT_APPROACH_AXLE", "pointApproachSection", section.get("id"))
    for section in line_map.get("cbtcApproachSections", []):
        for logical_id in section.get("logicalSectionIds", []):
            check_ref(logical_id, logical_ids, "CBTC_APPROACH_LOGICAL", "cbtcApproachSection", section.get("id"))
    for section in line_map.get("pointTriggerSections", []):
        for axle_id in section.get("axleSectionIds", []):
            check_ref(axle_id, axle_ids, "POINT_TRIGGER_AXLE", "pointTriggerSection", section.get("id"))
    for section in line_map.get("cbtcTriggerSections", []):
        for logical_id in section.get("logicalSectionIds", []):
            check_ref(logical_id, logical_ids, "CBTC_TRIGGER_LOGICAL", "cbtcTriggerSection", section.get("id"))

    for route in line_map.get("routes", []):
        entity_id = route.get("id")
        check_ref(route.get("startSignalId"), signal_ids, "ROUTE_SIGNAL", "route", entity_id)
        check_ref(route.get("endSignalId"), signal_ids, "ROUTE_SIGNAL", "route", entity_id)
        for axle_id in route.get("axleSectionIds", []):
            check_ref(axle_id, axle_ids, "ROUTE_AXLE", "route", entity_id)
        for section_id in route.get("protectionSectionIds", []):
            check_ref(section_id, protection_ids, "ROUTE_PROTECTION", "route", entity_id)
        for section_id in route.get("pointApproachSectionIds", []):
            check_ref(section_id, point_approach_ids, "ROUTE_POINT_APPROACH", "route", entity_id)
        for section_id in route.get("cbtcApproachSectionIds", []):
            check_ref(section_id, cbtc_approach_ids, "ROUTE_CBTC_APPROACH", "route", entity_id)
        for section_id in route.get("pointTriggerSectionIds", []):
            check_ref(section_id, point_trigger_ids, "ROUTE_POINT_TRIGGER", "route", entity_id)
        for section_id in route.get("cbtcTriggerSectionIds", []):
            check_ref(section_id, cbtc_trigger_ids, "ROUTE_CBTC_TRIGGER", "route", entity_id)

    report.ok = len(report.errors) == 0
    return report
